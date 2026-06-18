from django.shortcuts import render
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from collections import defaultdict

from .models import (
    Timetable,
    Program,
    Semester,
    Lecturer,
    CourseDifficultyPrediction
)

SLOTS = ["07:45", "09:45", "11:45", "13:45", "15:45", "17:45"]
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# =========================
# CONFLICT DETECTION
# =========================
def detect_conflicts(queryset):
    conflict_ids = set()

    data = list(queryset)

    for i in range(len(data)):
        for j in range(i + 1, len(data)):

            a = data[i]
            b = data[j]

            if a.day != b.day:
                continue

            # overlap check
            if a.start_time < b.end_time and a.end_time > b.start_time:

                if a.room_id == b.room_id or a.lecturer_id == b.lecturer_id:
                    conflict_ids.add(a.id)
                    conflict_ids.add(b.id)

    return conflict_ids
# =========================
# DASHBOARD
# =========================
def dashboard(request):
    return render(request, "dashboard.html", {
        "courses": Timetable.objects.values("course").distinct().count(),
        "lecturers": Lecturer.objects.count(),
        "slots": Timetable.objects.count(),
        "predictions": CourseDifficultyPrediction.objects.count()
    })


# =========================
# TIMETABLE VIEW (WITH CONFLICTS)
# =========================

from collections import defaultdict

def timetable_view(request):

    data = Timetable.objects.select_related(
        "course", "lecturer", "room", "semester"
    )

    # each cell can hold MULTIPLE classes
    grid = {slot: {day: [] for day in DAYS} for slot in SLOTS}

    for t in data:
        time_key = t.start_time.strftime("%H:%M")

        if time_key in grid:
            if t.day in grid[time_key]:
                grid[time_key][t.day].append(t)

    return render(request, "timetable.html", {
        "grid": grid,
        "days": DAYS,
        "slots": SLOTS,
        "programs": Program.objects.all(),
        "semesters": Semester.objects.all()
    })
# =========================
# LECTURER VIEW
# =========================
def lecturer_timetable(request, lecturer_id):
    data = Timetable.objects.filter(lecturer_id=lecturer_id)

    return render(request, "lecturer.html", {
        "timetable": data
    })


# =========================
# PDF EXPORT
# =========================
def export_timetable_pdf(request):
    data = Timetable.objects.select_related(
        "course", "lecturer", "room", "semester"
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="timetable.pdf"'

    p = canvas.Canvas(response)
    y = 800

    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, 820, "University Timetable")

    p.setFont("Helvetica", 10)

    for t in data:
        line = f"{t.course} | {t.lecturer} | {t.room} | {t.day} {t.start_time}-{t.end_time}"
        p.drawString(40, y, line)
        y -= 18

        if y < 50:
            p.showPage()
            y = 800

    p.save()
    return response


# =========================
# REST API Endpoints
# =========================
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.http import JsonResponse
import openpyxl
from io import BytesIO

from .serializers import CourseSerializer, TimetableSerializer, DifficultySerializer
from .services.scheduler import generate_timetable, DEFAULT_SLOTS


@api_view(["POST"])
def api_generate_timetable(request):
    payload = request.data
    program_code = payload.get("program_code")
    semester_name = payload.get("semester")

    try:
        program = Program.objects.get(code=program_code)
    except Program.DoesNotExist:
        return Response({"success": False, "message": "Program not found"}, status=404)

    semester = None
    if semester_name:
        semester = Semester.objects.filter(name=semester_name).first()
    if semester is None:
        semester = Semester.objects.first()

    generate_timetable(program, semester)

    return Response({"success": True, "message": "Timetable generated"})


@api_view(["GET"])
def api_preview_difficulties(request):
    program_code = request.GET.get("program_code")
    try:
        program = Program.objects.get(code=program_code)
    except Program.DoesNotExist:
        return Response({"success": False, "message": "Program not found"}, status=404)

    courses = Program.objects.get(code=program_code).course_set.all()
    result = []
    for c in courses:
        pred = CourseDifficultyPrediction.objects.filter(course=c).first()
        result.append({
            "code": c.code,
            "title": c.title,
            "difficulty_score": getattr(c, "difficulty_score", 0.0),
            "predicted_level": pred.predicted_level if pred else None,
            "confidence_score": pred.confidence_score if pred else None,
        })

    return Response({"success": True, "courses": result})


@api_view(["GET"])
def api_export_timetable_xlsx(request):
    # If program_code provided, export for that program only. Otherwise export for all programs.
    program_code = request.GET.get("program_code")
    all_programs = request.GET.get("all_programs") == "1" or program_code is None

    semester_name = request.GET.get("semester")
    semester = None
    if semester_name:
        semester = Semester.objects.filter(name=semester_name).first()
    if semester is None:
        semester = Semester.objects.first()

    programs = []
    if all_programs:
        programs = list(Program.objects.all())
    else:
        try:
            programs = [Program.objects.get(code=program_code)]
        except Program.DoesNotExist:
            return Response({"success": False, "message": "Program not found"}, status=404)

    # ensure timetables exist for each program by running the scheduler (respects constraints)
    for prog in programs:
        generate_timetable(prog, semester)

    wb = openpyxl.Workbook()
    # create sheets for Monday-Friday
    for day in DAYS:
        ws = wb.create_sheet(title=day)
        ws.cell(row=1, column=1, value=day.upper())

        # headers: start times in row 2, end times in row 3 (use DEFAULT_SLOTS if present else SLOTS)
        try:
            slots = DEFAULT_SLOTS
        except NameError:
            slots = [(s, s) for s in SLOTS]

        for i, pair in enumerate(slots, start=2):
            start, end = pair

            def format_am_pm(time_str):
                h, m = map(int, time_str.split(":"))
                suffix = "AM" if h < 12 else "PM"
                h12 = h % 12 or 12
                return f"{h12}:{m:02d}{suffix}"

            ws.cell(row=2, column=i, value=format_am_pm(start))
            ws.cell(row=3, column=i, value=format_am_pm(end))

        row = 4
        for prog in programs:
            ws.cell(row=row, column=1, value=prog.code or prog.name)

            for i, pair in enumerate(slots, start=2):
                start = pair[0] if isinstance(pair, (list, tuple)) else pair
                start_str = start
                matched = Timetable.objects.filter(
                    program=prog,
                    semester=semester,
                    day=day,
                    start_time=start_str,
                ).select_related("course", "room").first()

                if matched:
                    ws.cell(row=row, column=i, value=f"{matched.course.title}")
                    ws.cell(row=row + 1, column=i, value=f"{matched.room.name}")

            row += 2

    # remove default initial sheet if exists
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    fname = 'timetable_all_programs.xlsx' if all_programs else f'timetable_{programs[0].code}.xlsx'
    response = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
